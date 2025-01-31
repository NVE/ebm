import math
import pathlib
import typing

import numpy as np
from loguru import logger
import pandas as pd
from openpyxl.cell import Cell
from openpyxl.worksheet.errors import IgnoredError

from ebm.cmd.run_calculation import (calculate_building_category_area_forecast,
                                     calculate_building_category_energy_requirements,
                                     calculate_heating_systems)
from ebm.model.calibrate_heating_systems import transform_heating_systems
from ebm.model.building_condition import BEMA_ORDER as building_condition_order
from ebm.model.building_category import BEMA_ORDER as building_category_order
from ebm.model.data_classes import YearRange
from ebm.model.tek import BEMA_ORDER as tek_order


def transform_model_to_horizontal(model):
    hz = model.reset_index().copy()
    value_column = 'm2'
    if 'energy_requirement' in hz.columns:
        value_column = 'GWh'
        hz['GWh'] = hz['energy_requirement'] / 10**6
    hz = hz.groupby(by=['building_category', 'TEK', 'building_condition', 'year'], as_index=False).sum()[
        ['building_category', 'TEK', 'building_condition', 'year', value_column]]
    hz = hz.pivot(columns=['year'], index=['building_category', 'TEK', 'building_condition'], values=[
        value_column]).reset_index()

    hz = hz.sort_values(by=['building_category', 'TEK', 'building_condition'],
                        key=lambda x: x.map(building_category_order) if x.name == 'building_category' else x.map(
                            tek_order) if x.name == 'TEK' else x.map(
                            building_condition_order) if x.name == 'building_condition' else x)
    hz.insert(3, 'U', value_column)
    hz.columns = ['building_category', 'TEK', 'building_condition', 'U'] + [y for y in range(2020, 2051)]

    return hz

def transform_to_sorted_heating_systems(df: pd.DataFrame, holiday_homes: pd.DataFrame) -> pd.DataFrame:
    category_order = {'Bolig': 100, 'Fritidsboliger': 200, 'Yrkesbygg': 300}
    energy_source = {'Elektrisitet': 10, 'Fjernvarme': 11,  'Bio': 12, 'Fossil': 13, 'Solar': 13,
                     'Luft/luft': 24,  'Vannbåren varme': 25}

    rs = pd.concat([df, holiday_homes]).reindex()
    rs = rs.sort_values(by=['building_category', 'energy_source'],
                   key=lambda x: x.map(category_order) if x.name == 'building_category' else x.map(
                       energy_source) if x.name == 'energy_source' else x)

    hz = pd.concat([rs[~rs.energy_source.isin(['Luft/luft', 'Vannbåren varme'])],
                      rs[rs.energy_source.isin(['Luft/luft', 'Vannbåren varme'])]])
    hz.insert(2, 'U', 'GWh')
    return hz


def transform_holiday_homes_to_horizontal(df: pd.DataFrame) -> pd.DataFrame:
    df = df.reset_index()
    df = df.rename(columns={'energy_type': 'energy_source'})
    columns_to_keep = [y for y in YearRange(2020, 2050)] + ['building_category', 'energy_source']
    df = df.drop(columns=[c for c in df.columns if c not in columns_to_keep])
    df['energy_source'] = df['energy_source'].apply(lambda x: 'Elektrisitet' if x == 'electricity' else 'Bio' if x == 'fuelwood' else x)
    df['building_category'] = 'Fritidsboliger'
    return df

def transform_heating_systems_to_horizontal(model: pd.DataFrame):
    hs2 = model
    d = []
    for year in range(2020, 2051):
        energy_source_by_building_group = transform_heating_systems(hs2, year)
        energy_source_by_building_group['year'] = year
        d.append(energy_source_by_building_group)
    r = pd.concat(d)
    r2 = r.reset_index()[['building_category', 'energy_source', 'year', 'energy_use']]
    hz = r2.pivot(columns=['year'], index=['building_category', 'energy_source'],
                  values=['energy_use']).reset_index()

    hz.columns = ['building_category', 'energy_source'] + [y for y in range(2020, 2051)]
    return hz


def write_result(output_file, csv_delimiter, output, sheet_name='area forecast'):
    logger.debug(f'Writing to {output_file}')
    if str(output_file) == '-':
        try:
            print(output.to_markdown())

        except ImportError:
            print(output.to_string())
    elif output_file.suffix == '.csv':
        output.to_csv(output_file, sep=csv_delimiter)
        logger.info(f'Wrote {output_file}')
    else:
        excel_writer = pd.ExcelWriter(output_file, engine='openpyxl')
        output.to_excel(excel_writer, sheet_name=sheet_name, merge_cells=False, freeze_panes=(1, 3))
        excel_writer.close()
        logger.info(f'Wrote {output_file}')


def write_horizontal_excel(output_file: pathlib.Path, model: pd.DataFrame, sheet_name='Sheet 1'):
    more_options = {'mode': 'w'}
    if output_file.is_file():
        more_options = {'if_sheet_exists': 'replace', 'mode': 'a'}

    with pd.ExcelWriter(output_file, engine='openpyxl', **more_options) as writer:
        model.to_excel(writer, sheet_name=sheet_name, index=False)
        sheet = writer.sheets[sheet_name]

        columns = sheet.iter_cols(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column)

        logger.debug('Formatting cell')
        for col, (col_name, col_values) in zip(columns, model.items()):
            cell_format = detect_format_from_values(col_name, col_values, model)
            if cell_format:
                for row in col[1:]:
                    row.number_format = cell_format

        logger.debug('Adjust columns width')
        for col in sheet.iter_cols(min_col=1):
            adjusted_width = find_max_column_width(col)
            sheet.column_dimensions[col[0].column_letter].width = adjusted_width
        logger.debug(f'Closing {output_file} {sheet_name}')
    logger.debug(f'Wrote {output_file} {sheet_name}')


def detect_format_from_values(col_name, col_values, model):
    cell_format = ''
    if np.issubdtype(model[col_name].dtype, np.floating):
        cell_format = '#,##0.00'
        if col_values.max() > 1000.0:
            cell_format = '#,##0'
        elif 1.0 >= col_values.mean() >= -1.0:
            cell_format = '0.00%'
    elif np.issubdtype(model[col_name].dtype, np.integer):
        cell_format = '#,##0'

    return cell_format


def find_max_column_width(col: typing.Tuple[Cell]):
    max_length = 5
    for cell in col:
        try:
            if cell.value is not None:
                cell_length = len(str(cell.value)) + 1
                if cell.data_type == 'n':
                    thousands = math.floor(math.log(1_000_000_0000, 1000))
                    cell_length = max(len(str(cell.value).split('.')[0]) + thousands, 2)
                if cell_length > max_length:
                    max_length = cell_length
        except (AttributeError, KeyError, IgnoredError, ValueError) as ex:
            logger.debug(f'Got error f{cell.column_letter}')
            logger.error(ex)
            pass
    return max_length


def write_tqdm_result(output_file, output, csv_delimiter=','):
    try:
        from tqdm import tqdm
    except ImportError:
        # When tqdm is not installed we use write_result instead
        write_result(output_file, csv_delimiter, output)
        return

    logger.debug(f'Writing to {output_file}')
    if str(output_file) == '-':
        try:
            print(output.to_markdown())
        except ImportError:
            print(output.to_string())
            return

    output = output.reset_index()

    with tqdm(total=len(output), desc="Writing to spreadsheet") as pbar:
        if output_file.suffix == '.csv':
            for i in range(0, len(output), 100):  # Adjust the chunk size as needed
                building_category = output.iloc[i].building_category
                pbar.update(100)
                output.iloc[i:i + 100].to_csv(output_file, mode='a', header=(i == 0), index=False, sep=csv_delimiter)
                pbar.display(f'Writing {building_category}')
            pbar.display(f'Wrote {output_file}')
        else:
            with pd.ExcelWriter(output_file, engine='xlsxwriter') as excel_writer:
                for i in range(0, len(output), 100):  # Adjust the chunk size as needed
                    building_category = output.iloc[i].name[0] if 'building_category' not in output.columns else output.building_category.iloc[i]
                    pbar.set_description(f'Writing {building_category}')
                    start_row = 0 if i == 0 else i+1
                    page_start = i
                    page_end = min(i+100, len(output))
                    logger.trace(f'{start_row=} {page_start=} {page_end=}')
                    output.iloc[page_start:page_end].to_excel(excel_writer, startrow=start_row, header=(i == 0), merge_cells=False, index=False)
                    pbar.update(100)
                pbar.set_description(f'Closing {output_file}')


class EbmDefaultHandler:
    def extract_model(self, arguments, building_categories, database_manager, step_choice):

        area_forecast_results = []
        logger.debug('Extracting area forecast')
        for building_category in building_categories:
            area_forecast = calculate_building_category_area_forecast(
                building_category=building_category,
                database_manager=database_manager,
                start_year=arguments.start_year,
                end_year=arguments.end_year)
            area_forecast_results.append(area_forecast)

        df = pd.concat(area_forecast_results)

        df = df.set_index(['building_category', 'TEK', 'building_condition', 'year'])
        if 'energy-requirements' in step_choice or 'heating-systems' in step_choice:
            logger.debug('Extracting area energy requirements')
            energy_requirements_result = calculate_building_category_energy_requirements(
                building_category=building_categories,
                area_forecast=df,
                database_manager=database_manager,
                start_year=arguments.start_year,
                end_year=arguments.end_year)
            df = energy_requirements_result

            if 'heating-systems' in step_choice:
                logger.debug('Extracting heating systems')
                df = calculate_heating_systems(energy_requirements=energy_requirements_result,
                                               database_manager=database_manager)
        return df

    def write_result2(self, output_file, csv_delimiter, model):
        write_tqdm_result(output_file, model, csv_delimiter)
