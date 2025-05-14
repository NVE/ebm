import os
import pathlib

import dotenv
import pandas as pd
from dotenv import load_dotenv
from loguru import logger
from openpyxl.formatting.rule import FormulaRule
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from ebm.cmd.result_handler import transform_model_to_horizontal
from ebm.cmd.run_calculation import configure_loglevel
from ebm.energy_consumption import EnergyConsumption
from ebm.heating_systems_projection import HeatingSystemsProjection
from ebm.model.area_forecast import AreaForecast
from ebm.model.building_category import BuildingCategory
from ebm.model.building_condition import BuildingCondition
from ebm.model.buildings import Buildings
from ebm.model.construction import ConstructionCalculator
from ebm.model.data_classes import YearRange
from ebm.model.database_manager import DatabaseManager
from ebm.model.energy_requirement import EnergyRequirement
from ebm.model.file_handler import FileHandler
from ebm.model.building_category import BEMA_ORDER as building_category_order
from ebm.model.tek import BEMA_ORDER as tek_order
from ebm.model import heat_pump as h_p
from ebm.model import energy_use as e_u
from ebm.services.spreadsheet import find_max_column_width


def extract_energy_need(years: YearRange, dm: DatabaseManager) -> pd.DataFrame:
    er_calculator = EnergyRequirement.new_instance(period=years, calibration_year=2023,
                                                   database_manager=dm)
    energy_need = er_calculator.calculate_for_building_category(database_manager=dm)

    energy_need = energy_need.set_index(['building_category', 'TEK', 'purpose', 'building_condition', 'year'])

    return energy_need

def extract_area_change(years: YearRange, dm: DatabaseManager) -> tuple[pd.DataFrame, pd.DataFrame]:
    construction: pd.DataFrame | None = None
    demolition: pd.DataFrame | None = None

    tek_params = dm.get_tek_params(dm.get_tek_list())
    for building_category in BuildingCategory:
        buildings = Buildings.build_buildings(building_category=building_category, database_manager=dm, period=years)
        area_forecast: AreaForecast = buildings.build_area_forecast(dm, years.start, years.end)

        nested_list = [area_forecast.calc_area_pre_construction(t, BuildingCondition.DEMOLITION).to_frame().assign(**{'TEK': t}) for t in tek_params if t not in ['TEK17', 'TEK21']]
        bc_demolition = pd.concat(nested_list)
        bc_demolition['building_category'] = building_category

        demolition_floor_area = bc_demolition.groupby(by=['year']).sum().area
        df = ConstructionCalculator.calculate_construction(building_category, demolition_floor_area, dm, period=years)
        df['building_category'] = building_category
        df = df.reset_index().set_index(['building_category', 'year'])

        if construction is None:
            construction = df
            demolition = bc_demolition
        else:
            construction = pd.concat([construction, df])
            demolition = pd.concat([demolition, bc_demolition])

    return construction, demolition


def extract_area_forecast(years: YearRange, construction: pd.DataFrame, dm: DatabaseManager) -> pd.DataFrame:
    forecasts: pd.DataFrame | None = None

    for building_category in BuildingCategory:
        buildings = Buildings.build_buildings(building_category=building_category, database_manager=dm,
                                              period=years)
        area_forecast: AreaForecast = buildings.build_area_forecast(dm, years.start, years.end)

        accumulated_constructed_floor_area = construction.loc[building_category, 'accumulated_constructed_floor_area']
        forecast: pd.DataFrame = area_forecast.calc_area(accumulated_constructed_floor_area)
        forecast['building_category'] = building_category
        if forecasts is None:
            forecasts = forecast
        else:
            forecasts = pd.concat([forecasts, forecast])

    return forecasts


def transform_energy_need_to_energy_purpose_wide(energy_need, area_forecast):
    df_a = area_forecast.copy()
    df_a = df_a.query('building_condition!="demolition"').reset_index().set_index(
        ['building_category', 'building_condition', 'TEK', 'year'], drop=True)

    df_e = energy_need.copy().reset_index().set_index(
        ['building_category', 'building_condition', 'TEK', 'purpose', 'year'])

    df = df_e.join(df_a)[['m2', 'kwh_m2']].reset_index()
    df.loc[:, 'GWh'] = (df['m2'] * df['kwh_m2']) / 1_000_000
    df.loc[:, ('TEK', 'building_condition')] = ('all', 'all')

    non_residential = [b for b in BuildingCategory if b.is_non_residential()]

    df.loc[df[df['building_category'].isin(non_residential)].index, 'building_category'] = 'non_residential'

    df = df.groupby(by=['building_category', 'purpose', 'year'], as_index=False).sum()
    df = df[['building_category', 'purpose', 'year', 'GWh']]

    df = df.pivot(columns=['year'], index=['building_category', 'purpose'], values=['GWh']).reset_index()
    df = df.sort_values(by=['building_category', 'purpose'],
                        key=lambda x: x.map(building_category_order) if x.name == 'building_category' else x.map(
                            tek_order) if x.name == 'building_category' else x.map(
                            {'heating_rv': 1, 'heating_dhw': 2, 'fans_and_pumps': 3, 'lighting': 4,
                             'electrical_equipment': 5, 'cooling': 6}) if x.name == 'purpose' else x)

    df.insert(2, 'U', 'GWh')
    df.columns = ['building_category', 'purpose', 'U'] + [y for y in range(2020, 2051)]

    energy_purpose_fane1 = df

    return df


def transform_energy_need_to_energy_purpose_long(energy_need, area_forecast):
    df_a = area_forecast.copy()
    df_a = df_a.query('building_condition!="demolition"').reset_index().set_index(
        ['building_category', 'building_condition', 'TEK', 'year'], drop=True)

    df_e = energy_need.copy().reset_index().set_index(
        ['building_category', 'building_condition', 'TEK', 'purpose', 'year'])

    df = df_e.join(df_a)[['m2', 'kwh_m2']].reset_index()
    df.loc[:, 'GWh'] = (df['m2'] * df['kwh_m2']) / 1_000_000

    df = df.groupby(by=['year', 'building_category', 'TEK', 'purpose'], as_index=False).sum()
    df = df[['year', 'building_category', 'TEK', 'purpose', 'GWh']]
    df = df.sort_values(by=['year', 'building_category', 'TEK', 'purpose'],
                        key=lambda x: x.map(building_category_order) if x.name == 'building_category' else x.map(
                            tek_order) if x.name == 'building_category' else x.map(
                            tek_order) if x.name == 'TEK' else x.map(
                            {'heating_rv': 1, 'heating_dhw': 2, 'fans_and_pumps': 3, 'lighting': 4,
                             'electrical_equipment': 5, 'cooling': 6}) if x.name == 'purpose' else x)

    df = df.rename(columns={'GWh': 'energy_use [GWh]'})

    df.reset_index(inplace=True, drop=True)
    return df


def extract_heating_systems_projection(years, database_manager):
    projection_period = YearRange(2023, 2050)
    hsp = HeatingSystemsProjection.new_instance(projection_period, database_manager)
    df = hsp.calculate_projection()
    df = hsp.pad_projection(df, YearRange(2020, 2022))

    heating_system_projection = df.copy()
    return heating_system_projection


def extract_heating_systems(heating_system_projection, energy_need):
    calculator = EnergyConsumption(heating_system_projection.copy())

    calculator.heating_systems_parameters = calculator.grouped_heating_systems()

    df = calculator.calculate(energy_need)

    return df

    return heating_systems


def transform_heating_systems_share_long(heating_systems_projection):
    df = heating_systems_projection.copy()

    value_column = 'TEK_shares'

    fane2_columns = ['building_category', 'heating_systems', 'year', 'TEK_shares']

    df.loc[~df['building_category'].isin(['house', 'apartment_block']), 'building_category'] = 'non_residential'

    mean_tek_shares_yearly = df[fane2_columns].groupby(by=['year', 'building_category', 'heating_systems']).mean()
    return mean_tek_shares_yearly


def transform_heating_systems_share_wide(heating_systems_share_long):
    value_column = 'TEK_shares'
    df = heating_systems_share_long.copy().reset_index()
    df = df.pivot(columns=['year'], index=['building_category', 'heating_systems'], values=[value_column]).reset_index()

    df = df.sort_values(by=['building_category', 'heating_systems'],
                        key=lambda x: x.map(building_category_order) if x.name == 'building_category' else x)
    df.insert(2, 'U', value_column)
    df['U'] = '%'

    df.columns = ['building_category', 'heating_systems', 'U'] + [y for y in range(2020, 2051)]
    return df


def heating_systems_parameter_from_projection(heating_systems_projection):
    calculator = EnergyConsumption(heating_systems_projection.copy())

    return calculator.grouped_heating_systems()


def extract_energy_use_kwh(heating_systems_parameter, energy_need):
    df = e_u.all_purposes(heating_systems_parameter)
    df.loc[:, 'building_group'] = 'yrkesbygg'
    df.loc[df.building_category.isin(['house', 'apartment_block']), 'building_group'] = 'bolig'

    efficiency_factor = e_u.efficiency_factor(df)
    energy_use_kwh = e_u.energy_use_kwh(energy_need=energy_need, efficiency_factor=efficiency_factor)

    return energy_use_kwh


def transform_demolition_construction(energy_need, demolition, construction):
    demolition['demolition_construction'] = 'demolition'
    demolition['energy_use'] = demolition['area']
    demolition['m2'] = -demolition['area']

    construction['demolition_construction'] = 'construction'
    construction['TEK'] = 'TEK17'
    construction['energy_use'] = -construction['constructed_floor_area']
    construction['m2'] = construction['constructed_floor_area']

    return pd.concat([
        demolition.reset_index()[['year', 'demolition_construction', 'building_category', 'TEK', 'm2', 'energy_use']],
        construction.reset_index()[['year', 'demolition_construction', 'building_category', 'TEK', 'm2', 'energy_use']]
    ])


def main():
    env_file = pathlib.Path(dotenv.find_dotenv(usecwd=True))
    if env_file.is_file():
        logger.debug(f'Loading environment from {env_file}')
        load_dotenv(pathlib.Path('.env').absolute())
    else:
        logger.debug(f'.env not found in {env_file.absolute()}')

    configure_loglevel(os.environ.get('LOG_FORMAT', None))

    years = YearRange(2020, 2050)
    input_path = pathlib.Path('kalibar')
    output_path = pathlib.Path('t2775_output')
    output_path.mkdir(exist_ok=True)

    file_handler = FileHandler(directory=input_path)
    database_manager = DatabaseManager(file_handler=file_handler)

    logger.info('✅ Area to area.xlsx')

    logger.debug('✅ extract area_change')
    construction, demolition = extract_area_change(years, database_manager)
    logger.debug('✅ extract area')
    forecasts = extract_area_forecast(years, construction, database_manager)

    logger.debug('✅ transform fane 1 (wide)')

    df = forecasts.copy()

    df = df.query('building_condition!="demolition"')
    df.loc[:, 'TEK'] = 'all'
    df.loc[:, 'building_condition'] = 'all'

    df = transform_model_to_horizontal(df).drop(columns=['TEK', 'building_condition'])

    area_fane_1 = df.copy()

    logger.debug('✅ transform fane 2 (long')

    df = forecasts['year,building_category,TEK,building_condition,m2'.split(',')].copy()
    df = df.query('building_condition!="demolition"')

    df = df.groupby(by='year,building_category,TEK'.split(','))[['m2']].sum().rename(columns={'m2': 'area'})
    df.insert(0, 'U', 'm2')
    area_fane_2 = df.reset_index()

    logger.debug('✅ Write file area.xlsx')

    area_output = output_path / 'area.xlsx'

    with pd.ExcelWriter(area_output, engine='xlsxwriter') as writer:
        logger.debug('✅ reorder columns')
        logger.debug('✅ make area.xlsx pretty')
        area_fane_1.to_excel(writer, sheet_name='wide', index=False)
        area_fane_2.to_excel(writer, sheet_name='long', index=False)
    make_pretty(area_output)


    logger.info('✅ Energy use to energy_purpose')
    logger.debug('✅ extract energy_need')
    energy_need = extract_energy_need(years, database_manager)

    total_energy_need = forecasts.reset_index().set_index(['building_category', 'TEK', 'building_condition', 'year']).merge(energy_need, left_index=True, right_index=True)
    total_energy_need['energy_requirement'] = total_energy_need.kwh_m2 * total_energy_need.m2

    logger.debug('✅ transform fane 1')
    energy_purpose_fane1 = transform_energy_need_to_energy_purpose_wide(energy_need=energy_need, area_forecast=forecasts)
    logger.debug('✅ transform fane 2')
    energy_purpose_fane2 = transform_energy_need_to_energy_purpose_long(energy_need=energy_need, area_forecast=forecasts)

    logger.debug('✅ Write file energy_purpose.xlsx')
    energy_purpose_output = output_path / 'energy_purpose.xlsx'

    with pd.ExcelWriter(energy_purpose_output, engine='xlsxwriter') as writer:
        logger.debug('❌ reorder columns')
        logger.debug(f'❌ make {energy_purpose_output.name} pretty')
        energy_purpose_fane1.to_excel(writer, sheet_name='wide')
        energy_purpose_fane2.to_excel(writer, sheet_name='long')



    logger.info('✅ Heating_system_share')

    heating_systems_projection = extract_heating_systems_projection(years, database_manager)
    logger.debug('✅ transform fane 2')
    heating_systems_share = transform_heating_systems_share_long(heating_systems_projection)
    logger.debug('✅ transform fane 1')
    heating_systems_share_wide = transform_heating_systems_share_wide(heating_systems_share)

    logger.debug('✅ Write file heating_system_share.xlsx')
    heating_system_share = output_path / 'heating_system_share.xlsx'

    with pd.ExcelWriter(heating_system_share, engine='xlsxwriter') as writer:
        logger.debug('❌ reorder columns')
        logger.debug(f'❌ make {heating_system_share.name} pretty')
        heating_systems_share_wide.to_excel(writer, sheet_name='wide')
        heating_systems_share.to_excel(writer, sheet_name='long')

    logger.info('✅ heat_prod_hp')

    logger.debug('✅ extract heating_system_parameters')
    heating_systems_parameter = heating_systems_parameter_from_projection(heating_systems_projection)
    logger.debug('✅ transform to hp')

    df = heating_systems_parameter
    df = df.assign(**{'heating_system': df['heating_systems'].str.split('-')}).explode('heating_system')
    df['heating_system'] = df['heating_system'].str.strip()
    df['load_share'] = df['Grunnlast andel']

    air_air = h_p.air_source_heat_pump(df)
    district_heating = h_p.district_heating_heat_pump(df)
    production = h_p.heat_pump_production(total_energy_need, air_air, district_heating)

    heat_prod_hp_wide = h_p.heat_prod_hp_wide(production)

    logger.debug('✅ Write file heat_prod_hp.xlsx')
    heat_prod_hp_file = output_path / 'heat_prod_hp.xlsx'

    with pd.ExcelWriter(heat_prod_hp_file, engine='xlsxwriter') as writer:
        logger.debug('❌ reorder columns')
        logger.debug(f'❌ make {heat_prod_hp_file.name} pretty')
        heat_prod_hp_wide.to_excel(writer, sheet_name='wide')

    logger.info('❌ Energy_use')

    logger.debug('✅ extract energy_use_kwh')
    energy_use_kwh = extract_energy_use_kwh(heating_systems_parameter, total_energy_need)

    logger.debug('✅ transform fane 2')
    logger.debug('✅ group by category, year, product')
    e_u_by_category = energy_use_kwh[['building_category', 'year', 'energy_product', 'kwh']].groupby(
        by=['building_category', 'energy_product', 'year']).sum() / 1_000_000
    energy_use_long = e_u_by_category.rename(columns={'kwh': 'energy_use'})

    logger.debug('❌ transform fane 1')
    logger.debug('❌ add holiday homes 1')

    logger.debug('✅ group by group, product year')
    e_u_by_group = energy_use_kwh[['building_group', 'year', 'energy_product', 'kwh']].groupby(
        by=['building_group', 'energy_product', 'year']).sum() / 1_000_000
    energy_use_wide = e_u_by_group.reset_index().pivot(columns=['year'], index=['building_group', 'energy_product'], values=['kwh']).reset_index()

    logger.debug('✅ Write file energy_use')

    energy_use_file = output_path / 'energy_use.xlsx'

    with pd.ExcelWriter(energy_use_file, engine='xlsxwriter') as writer:
        logger.debug('❌ reorder columns')
        logger.debug(f'❌ make {energy_use_file.name} pretty')
        energy_use_long.to_excel(writer, sheet_name='long')
        energy_use_wide.to_excel(writer, sheet_name='wide')

    logger.info('Returns…')

    logger.debug('❌ Write file demolition_construction')
    logger.info('❌ building razing to demolition_construction.xlsx')


    demolition_construction = transform_demolition_construction(energy_need, demolition, construction)

    logger.debug('✅ extract demolition')
    logger.debug('✅ extract construction')
    logger.debug('❌ extract energy_need')
    logger.debug('✅ transform demolition_construction')

    demolition_construction_file = output_path / 'demolition_construction.xlsx'

    logger.debug('✅ Write file demolition_construction.xlsx')
    with pd.ExcelWriter(demolition_construction_file, engine='xlsxwriter') as writer:
        logger.debug('❌ reorder columns')
        logger.debug(f'❌ make {demolition_construction_file.name} pretty')
        demolition_construction.to_excel(writer, sheet_name='long')

    logger.info('❌ Ekstra resultater som skrives etter behov')
    logger.debug('❌ area')
    logger.debug('❌ energy_need')
    logger.debug('❌ energy_use')


def make_pretty(workbook_name):
    wb = load_workbook(workbook_name)
    font = Font(name='Arial', size=11)
    header_font = Font(name='Source Sans Pro', size=11, bold=True, color="ffffff")
    for s in wb.sheetnames:
        ws = wb[s]
        # Freeze the top row
        ws.freeze_panes = ws['A2']

        ws.font = font

        # Define the fill color
        header_fill = PatternFill(start_color='c8102e', end_color='c8102e', fill_type='solid')
        odd_fill = PatternFill(start_color='ffd8de', end_color='ffd8de', fill_type='solid')
        even_fill = PatternFill(start_color='ffebee', end_color='ffebee', fill_type='solid')

        # Apply the fill color to the header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row_number, row in enumerate(ws.rows):
            if row_number == 0:
                continue
            for column_number, cell in enumerate(row):
                cell.font = font

        even_rule = FormulaRule(formula=['MOD(ROW(),2)=0'], fill=even_fill)
        odd_rule = FormulaRule(formula=['MOD(ROW(),2)=1'], fill=odd_fill)
        worksheet_range = f'A2:{get_column_letter(ws.max_column)}{ws.max_row}'
        # logger.error(worksheet_range)
        ws.conditional_formatting.add(worksheet_range, odd_rule)
        ws.conditional_formatting.add(worksheet_range, even_rule)

        for col in ws.iter_cols(min_col=1):
            adjusted_width = find_max_column_width(col)
            ws.column_dimensions[col[0].column_letter].width = adjusted_width
            values = [int(r.value) for r in col if r.value and r.data_type == 'n']
            if values:
                max_value = max(values)
                if max_value > 1000:
                    for row_number, cell in enumerate(col):
                        if row_number < 1:
                            if cell.value == 'year':
                                break
                            continue
                        cell.number_format = '### ### ### ### ##0'
    wb.save(workbook_name)


if __name__ == '__main__':
    main()

