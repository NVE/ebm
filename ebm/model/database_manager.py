import pathlib
import typing

import pandas as pd
from openpyxl import load_workbook

from .file_handler import FileHandler
from .data_classes import ScurveParameters, TEKParameters
from ..services.spreadsheet import iter_cells


class DatabaseManager():
    """
    Manages database operations.
    """

    # TODO:
    # - change all strings to lower case and underscore instead of space

    # Column names
    COL_TEK = 'TEK'
    COL_TEK_BUILDING_YEAR = 'building_year'
    COL_TEK_START_YEAR = 'period_start_year'
    COL_TEK_END_YEAR = 'period_end_year'
    COL_BUILDING_CATEGORY = 'building_category'
    COL_BUILDING_CONDITION = 'condition'
    COL_EARLIEST_AGE = 'earliest_age_for_measure'
    COL_AVERAGE_AGE = 'average_age_for_measure'
    COL_LAST_AGE = 'last_age_for_measure'
    COL_RUSH_YEARS = 'rush_period_years'
    COL_RUSH_SHARE = 'rush_share'
    COL_NEVER_SHARE = 'never_share'
    
    def __init__(self):
        self.file_handler = FileHandler()

    def get_building_category_list(self):
        """
        Get a list of building categories.

        Returns:
        - building_category_list (list): List of building categories.
        """
        building_categories = self.file_handler.get_building_categories()
        building_category_list = building_categories[self.COL_BUILDING_CATEGORY].unique()
        return building_category_list

    def get_condition_list(self):
        """
        Get a list of building conditions.

        Returns:
        - condition_list (list): List of building conditions.
        """
        building_conditions = self.file_handler.get_building_conditions()
        condition_list = building_conditions[self.COL_BUILDING_CONDITION].unique()
        return condition_list
    
    def get_tek_list(self):
        """
        Get a list of TEK IDs.

        Returns:
        - tek_list (list): List of TEK IDs.
        """
        tek_id = self.file_handler.get_tek_id()
        tek_list = tek_id[self.COL_TEK].unique()
        return tek_list
    
    def get_tek_params(self, tek_list: typing.List[str]):
        """
        Retrieve TEK parameters for a list of TEK IDs.

        This method fetches TEK parameters for each TEK ID in the provided list,
        converts the relevant data to a dictionary, and maps these values to the 
        corresponding attributes of the TEKParameters dataclass. The resulting 
        dataclass instances are stored in a dictionary with TEK IDs as keys.

        Parameters:
        - tek_list (list of str): List of TEK IDs.

        Returns:
        - tek_params (dict): Dictionary where each key is a TEK ID and each value 
                            is a TEKParameters dataclass instance containing the 
                            parameters for that TEK ID.
        """
        tek_params = {}

        tek_params_df = self.file_handler.get_tek_params()

        for tek in tek_list:
            # Filter on TEK ID
            tek_params_filtered = tek_params_df[tek_params_df[self.COL_TEK] == tek]

            # Assuming there is only one row in the filtered DataFrame
            tek_params_row = tek_params_filtered.iloc[0]

            # Convert the single row to a dictionary
            tek_params_dict = tek_params_row.to_dict()

            # Map the dictionary values to the dataclass attributes
            tek_params_per_id = TEKParameters(
                tek = tek_params_dict[self.COL_TEK],
                building_year = tek_params_dict[self.COL_TEK_BUILDING_YEAR],
                start_year = tek_params_dict[self.COL_TEK_START_YEAR],
                end_year = tek_params_dict[self.COL_TEK_END_YEAR],
                )

            tek_params[tek] = tek_params_per_id    

        return tek_params

    def get_s_curve_params(self):
        """
        Get input dataframe with S-curve parameters/assumptions.

        Returns:
        - s_curve_params (pd.DataFrame): DataFrame with S-curve parameters.
        """
        s_curve_params = self.file_handler.get_s_curve_params()
        return s_curve_params

    def get_s_curve_params_per_building_category_and_condition(self, building_category: str, 
                                                               condition: str) -> ScurveParameters:
        """
        Get input dataframe with S-curve parameters/assumptions and filter it by building category and condition.

        Parameters:
        - building_category (str): Building category.
        - condition (str): Building condition.

        Returns:
        - s_curve_params_dict (dict): Dictionary containing S-curve parameters with column names as keys and 
                                      corresponding column values as values.
        """
        # Retrieve input data and filter on building category and condition
        s_curve_params = self.file_handler.get_s_curve_params()
        s_curve_params_filtered = s_curve_params[(s_curve_params[self.COL_BUILDING_CATEGORY] == building_category) & (s_curve_params[self.COL_BUILDING_CONDITION] == condition)]

        # Assuming there is only one row in the filtered DataFrame
        s_curve_params_row = s_curve_params_filtered.iloc[0]

        # Convert the single row to a dictionary
        s_curve_params_dict = s_curve_params_row.to_dict()
        
        # Map the dictionary values to the dataclass attributes
        scurve_parameters = ScurveParameters(
            building_category=s_curve_params_dict[self.COL_BUILDING_CATEGORY],
            condition=s_curve_params_dict[self.COL_BUILDING_CONDITION],
            earliest_age=s_curve_params_dict[self.COL_EARLIEST_AGE],
            average_age=s_curve_params_dict[self.COL_AVERAGE_AGE], 
            rush_years=s_curve_params_dict[self.COL_RUSH_YEARS], 
            last_age=s_curve_params_dict[self.COL_LAST_AGE],
            rush_share=s_curve_params_dict[self.COL_RUSH_SHARE],
            never_share=s_curve_params_dict[self.COL_NEVER_SHARE],
        )
        
        return scurve_parameters 

    def get_construction_population(self) -> pd.DataFrame:
        """
        Get construction population DataFrame.

        Returns:
        - construction_population (pd.DataFrame): Dataframe containing population numbers
          year population household_size
        """
        new_buildings_population = self.file_handler.get_construction_population()
        new_buildings_population["household_size"] = new_buildings_population['household_size'].astype('float64')
        new_buildings_population = new_buildings_population.set_index('year')
        return new_buildings_population

    def get_new_buildings_category_share(self) -> pd.DataFrame:
        """
        Get building category share by year as a DataFrame.

        The number can be used in conjunction with number of households to calculate total number
        of buildings of category house and apartment block

        Returns:
        - new_buildings_category_share (pd.DataFrame): Dataframe containing population numbers
          "year", "Andel nye småhus", "Andel nye leiligheter", "Areal nye småhus", "Areal nye leiligheter"
        """
        df = self.file_handler.get_construction_building_category_share()
        df['new_house_share'] = df['new_house_share'].astype('float64')
        df['new_apartment_block_share'] = df['new_apartment_block_share'].astype('float64')

        return df.set_index('year')

    def get_building_category_floor_area(self) -> pd.DataFrame:
        """
        Get population and household size DataFrame from a file.

        Returns:
        - construction_population (pd.DataFrame): Dataframe containing population numbers
          "area","type of building","2010","2011"
        """

        df = self.file_handler.get_building_category_area()
        types = df['type_of_building'].apply(lambda r: pd.Series(r.split(' ', 1)))
        df[['type_no', 'typename']] = types

        df = df.drop(columns=['type_of_building', 'area'])
        return df

    def get_building_category_area_by_tek(self, building_category=None) -> pd.DataFrame:
        """
        get total area of building_category by TEK.
        Parameters:
        - building_category (str): optional parameter that filter the returned dataframe by building_category
        Returns:
        - building_category_area_by_tek (pd.DataFrame): Dataframe containing area numbers
          "building_category","TEK","area"
        """

        df = self.file_handler.get_building_category_area_by_tek()
        df.loc[df.TEK == '-> TEK49', 'TEK'] = 'PRE_TEK49'
        if building_category:
            return df[df.building_category == building_category]
        return df

    def load_demolition_floor_area_from_spreadsheet(self, path: pathlib.Path = None, row: int = 656) -> pd.DataFrame:
        p = pathlib.Path("st_bema2019_a_hus.xlsx")
        if path:
            p = path
        wb = load_workbook(filename=p)
        sheet_ranges = wb.sheetnames
        sheet = wb[sheet_ranges[0]]
        cells = list(iter_cells(first_column='E'))[:41]

        a_hus_revet = pd.DataFrame({'demolition': [sheet[f'{c}{row}'].value for c in cells]}, index=list(range(2010, 2051)))
        #a_hus_revet = pd.DataFrame({'demolition': [sheet[f'{c}655'].value for c in cells]}, index=list(range(2010, 2051)))
        a_hus_revet['demolition'] = a_hus_revet['demolition']
        a_hus_revet['demolition_change'] = a_hus_revet.demolition.diff(1).fillna(0)
        a_hus_revet.index = a_hus_revet.index.rename('year')
        print(a_hus_revet)
        return a_hus_revet
