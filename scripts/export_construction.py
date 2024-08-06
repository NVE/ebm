import itertools

from openpyxl import load_workbook
from loguru import logger
import pandas as pd
import rich.pretty
from model.bema import load_construction_building_category, BuildingCategory


#spreadsheet = 'C:/Users/kenord/OneDrive - Norges vassdrags- og energidirektorat/Dokumenter/regneark/statiskBEMA_2019.xlsm'
spreadsheet = 'C:/Users/kenord/OneDrive - Norges vassdrags- og energidirektorat/Dokumenter/regneark/st_bema2019_nybygging.xlsx'


def main():
    wb = load_workbook(spreadsheet)
    worksheet_name = 'Nybygging' if 'Nybygging' in wb.sheetnames else wb.sheetnames[0]
    logger.debug(f'Opening {spreadsheet} {worksheet_name}')
    sheet = wb[worksheet_name]  # wb['Nybygging']  # Nybygging

    for building_category in list(BuildingCategory):
        logger.debug(building_category)
        df = load_construction_building_category(sheet, building_category)
        print('===', '  '.join(building_category.name.split()), '===')
        rich.pretty.pprint(df.transpose())
    wb.close()


if __name__ == '__main__':
    main()


