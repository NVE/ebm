import argparse
import itertools
import os

from openpyxl import load_workbook
from loguru import logger
import rich.pretty
from dotenv import load_dotenv

from model.bema import load_construction_building_category, BuildingCategory


def main():
    load_dotenv()

    arg_parser = argparse.ArgumentParser()
    arg_parser.add_argument('spreadsheet', type=str, nargs='?', default=os.environ.get('BEMA_SPREADSHEET'))
    arg_parser.add_argument('worksheet', type=str, nargs='?', default=None)

    arguments = arg_parser.parse_args()
    spreadsheet_name = arguments.spreadsheet
    if not spreadsheet_name:
        raise ValueError('Please provide a filename argument or use the environment variable BEMA_SPREADSHEET')

    wb = load_workbook(spreadsheet_name)
    worksheet_name = arguments.worksheet
    if not worksheet_name:
        if 'Nybygging' in wb.sheetnames:
            worksheet_name = 'Nybygging'
        else:
            worksheet_name = wb.sheetnames[0]

    logger.debug(f'Opening {spreadsheet_name} {worksheet_name}')
    sheet = wb[worksheet_name]

    building_categories = list(BuildingCategory)
    for building_category in building_categories:
        logger.debug(building_category)
        df = load_construction_building_category(sheet, building_category)
        transposed = df.transpose()

        print('===', '  '.join(building_category.name.split()), '===')
        rich.pretty.pprint(transposed)

    wb.close()


if __name__ == '__main__':
    main()


