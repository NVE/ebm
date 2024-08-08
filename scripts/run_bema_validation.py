import typing
import os

from openpyxl import load_workbook
from loguru import logger
from dotenv import load_dotenv
import pandas as pd


from ebm.model.building_category import BuildingCategory
from ebm.model.building_condition import BuildingCondition

## TESTING

#logger.debug('debug')
#logger.info('info')
#logger.warning('warning')
#logger.error('error')

folder_path = 'X:\Brukere-felles\lfep\BEMA'
file_name = 'test_file.xlsx'
workbook_name = os.path.join(folder_path, file_name)
worksheet_name = 'Sheet1'
#row =  

logger.debug(f'{workbook_name}')

workbook = load_workbook(workbook_name)
worksheet = workbook[worksheet_name]
cell_rows = worksheet.cell(row=1, column=1).value

print(cell_rows)


df = pd.read_excel(workbook_name, 
                   sheet_name=worksheet_name,
                   header=0,                   # number of rows to skip before header, e.g. header= 7, then row 8 is used as header 
                   #usecols="D:U",              # cols to include
                   #nrows=324                   # number of rows to include from header row
                   )

print(df)