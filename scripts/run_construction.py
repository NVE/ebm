import argparse
import collections.abc
import os
import sys
import typing

import pandas as pd
from dotenv import load_dotenv
from loguru import logger
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from rich.pretty import pprint

from ebm.model import BuildingCategory
from ebm.model.bema import load_construction_building_category
from ebm.model.construction import ConstructionCalculator
from model import DatabaseManager, Buildings

ROUND_PRECISION = 4

logger.info = pprint


def main():
    """ Program used to test nybygging calculation main function """
    load_dotenv()

    default_building_categories: typing.List[str] = [str(b) for b in iter(BuildingCategory)]

    arg_parser = argparse.ArgumentParser()
    arg_parser.add_argument('--debug', action='store_true')

    arg_parser.add_argument('building_categories', type=str, nargs='*', default=default_building_categories)
    arguments: argparse.Namespace = arg_parser.parse_args()

    if not arguments.debug and os.environ.get('DEBUG', '') != 'True':
        logger.remove()
        logger.add(sys.stderr, level="INFO")

    building_categories = arguments.building_categories

    error_count = validate_building_categories(building_categories)
    sys.exit(error_count)


def validate_building_categories(building_categories: collections.abc.Iterable[BuildingCategory | str]):
    """
    Validate building categories accumulated constructed floor area using the function
    validate_accumulated_constructed_floor_area()

    Parameters
    ----------
    building_categories Iterable of BuildingCategory that needs validation

    Returns error_count total number of errors found
    -------

    """
    error_count = 0
    categories_with_errors = []
    for b_category in building_categories:
        building_category: BuildingCategory = BuildingCategory.from_string(b_category)
        logger.info(f'Checking {building_category=}')

        category_error_count = validate_accumulated_constructed_floor_area(building_category)
        if category_error_count > 0:
            logger.error(f'Got {category_error_count} errors for category {building_category}')
            categories_with_errors.append((building_category, category_error_count))
        error_count = category_error_count + error_count
    if error_count > 0:
        logger.error(f'Got {error_count} errors in total')
        logger.error(f'{categories_with_errors=}')
    return error_count


def load_accumulated_constructed_floor_area(building_category: BuildingCategory) -> pd.Series:
    spreadsheet_name = os.environ.get('BEMA_SPREADSHEET')

    wb: Workbook = load_workbook(spreadsheet_name)
    sheet_name: Worksheet = wb[wb.sheetnames[0]]

    df = load_construction_building_category(worksheet=sheet_name, building_category=building_category)

    return df.accumulated_constructed_floor_area


def validate_accumulated_constructed_floor_area(building_category) -> int:
    expected_values: pd.Series = load_accumulated_constructed_floor_area(building_category)

    # Prerequisites for calculate_construction
    database_manager = DatabaseManager()
    demolition_floor_area = extract_demolition_floor_area(building_category, database_manager)

    yearly_constructed = ConstructionCalculator.calculate_construction(building_category, demolition_floor_area, database_manager)

    constructed_floor_area = yearly_constructed.accumulated_constructed_floor_area
    years = constructed_floor_area.index

    error_count = 0
    for year, current_value, expected_value in zip(years, constructed_floor_area, expected_values):
        if round(current_value, ROUND_PRECISION) != round(expected_value, ROUND_PRECISION):
            error_count = error_count + 1
            logger.error(f'Expected {expected_value} got {current_value} {building_category.name} {year=} ')
            logger.debug(f'The difference is {expected_value - current_value}. {ROUND_PRECISION=}')
            logger.debug(f'{yearly_constructed.loc[year]}')
        else:
            logger.debug(f'Found expected value {expected_value} {year=} {current_value=}')
    return error_count


def extract_demolition_floor_area(building_category: BuildingCategory,
                                  database_manager: DatabaseManager = None) -> pd.Series:
    """
    Extracts demolition_floor_area calculated from AreaForecast.

    Args:
        building_category: building category used for AreaForecast and Building i.e `kindergarten`
        database_manager: common DatabaseManager

    Returns: demolition_floor_area: pd.Series
        with years as the index
    """

    db = database_manager if database_manager else DatabaseManager()

    buildings = Buildings.build_buildings(building_category, db)
    area_forecast = buildings.build_area_forecast(db, start_year=2010, end_year=2050)

    years = [y for y in range(2010, 2050 + 1)]
    demolition_floor_area = pd.Series(data=area_forecast.calc_total_demolition_area_per_year(),
                                      index=years)
    return demolition_floor_area


if __name__ == '__main__':
    main()
