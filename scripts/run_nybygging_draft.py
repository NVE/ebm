#!/usr/bin/env python
# coding: utf-8

import itertools
import pathlib
import string
import typing

from IPython.display import display
from loguru import logger
from openpyxl import load_workbook
import numpy as np
import pandas as pd

from ebm.model import DatabaseManager


def calculate_households(population: pd.DataFrame) -> pd.DataFrame:
    population['households'] = (population['population'] / population['household_size'])
    population['population_change'] = population.population.diff(1) / population.population.shift(1)
    population['households_change'] = population.households.diff(1)
    population['building_change'] = population.households
    return population


def calculate_house_change(population: pd.DataFrame, new_buildings_category_share: pd.DataFrame) -> pd.DataFrame:
    floor_area = population.merge(new_buildings_category_share, left_on='year', right_on='year')
    floor_area['yearly_change_house_count'] = (
        (floor_area['households_change'] * floor_area['new_house_share'])).fillna(0)

    floor_area = floor_area[['yearly_change_house_count', 'floor_area_new_house', 'new_house_share']]

    floor_area['yearly_change_floor_area_house'] = (
        (floor_area['yearly_change_house_count'] *
         floor_area['floor_area_new_house'])).fillna(0)
    floor_area = floor_area[
        ['yearly_change_house_count', 'yearly_change_floor_area_house', 'floor_area_new_house']]
    if 'year' in floor_area.columns:
        floor_area = floor_area.set_index('year')
    return floor_area


def calculate_house_floor_area_demolished_by_year() -> pd.DataFrame:
    p = pathlib.Path(
        "C:/Users/kenord/OneDrive - Norges vassdrags- og energidirektorat/Dokumenter/regneark/st_bema2019_a_hus.xlsx")
    wb = load_workbook(filename=p)
    sheet_ranges = wb.sheetnames
    sheet = wb[sheet_ranges[0]]
    cells = list(iter_cells(first_column='E'))[:41]
    logger.info(list(string.ascii_uppercase)[4:] + [a + b for a, b in itertools.product(string.ascii_uppercase, repeat=2)][
                                               :19])
    logger.info(cells)
    a_hus_revet = pd.DataFrame({'demolition': [sheet[f'{c}656'].value for c in cells]}, index=list(range(2010, 2051)))
    a_hus_revet['demolition'] = a_hus_revet['demolition']
    a_hus_revet['demolition_change'] = a_hus_revet.demolition.diff(1).fillna(0)
    a_hus_revet.index = a_hus_revet.index.rename('year')

    return a_hus_revet


def calculate_yearly_built_floor_area_house(build_area: pd.DataFrame,
                                            floor_area_demolished_by_year: pd.DataFrame,
                                            yearly_change_small_house: pd.DataFrame) -> pd.DataFrame:
    build_area['building_category'] = 'unknown'
    build_area.loc[(build_area.type_no >= '111') & (build_area.type_no <= '136'), 'building_category'] = 'Small house'
    build_area.loc[
        (build_area.type_no >= '141') & (build_area.type_no <= '146'), 'building_category'] = 'Apartment block'
    build_area_sum = build_area.groupby(by='building_category').sum()

    yearly_change_small_house['yearly_new_house_floor_area'] = \
        yearly_change_small_house['yearly_change_floor_area_house'] + floor_area_demolished_by_year['demolition_change']
    yearly_change_small_house.loc[2010, 'yearly_new_house_floor_area'] = build_area_sum.loc['Small house']['2010']
    yearly_change_small_house.loc[2011, 'yearly_new_house_floor_area'] = build_area_sum.loc['Small house']['2011']
    return yearly_change_small_house


def calculate_new_buildings_accumulated(yearly_change_small_house) -> pd.DataFrame:
    yearly_change_small_house['new_house_accumulated'] = yearly_change_small_house[
        'yearly_new_house_floor_area'].cumsum()
    # In the Bema_2019 spreadsheet the value for 2010 is empty. Using zero instead. This will prevent NaN issues later.
    yearly_change_small_house.loc[2010, 'new_house_accumulated'] = 0
    return yearly_change_small_house


def iter_cells(first_column: str = 'E', left_padding: str = '') -> typing.Generator[str, None, None]:
    """ Returns spreadsheet column names from A up to ZZ
        Parameters:
        - first_column Letter of the first column to return. Default (E)
        - left_padding Padding added in front of single letter columns. Default empty
        Returns:
        - Generator supplying a column name
    """
    if not first_column:
        first_index = 0
    elif first_column.upper() not in string.ascii_uppercase:
        raise ValueError(f'Expected first_column {first_column} in {string.ascii_uppercase}')
    elif len(first_column) != 1:
        raise ValueError(f'Expected first_column of length 1 was: {len(first_column)}')
    else:
        first_index = string.ascii_uppercase.index(first_column.upper())
    for cell in string.ascii_uppercase[first_index:]:
        yield f'{left_padding}{cell}'
    for a, b in itertools.product(string.ascii_uppercase, repeat=2):
        yield a+b


def main():
    process_new_buildings_house()


def process_new_buildings_house():
    database_manager = DatabaseManager()
    population = database_manager.get_construction_population()
    households = calculate_households(population=population)

    facts = households[['population', 'population_change', 'household_size', 'households', 'households_change']]

    demolition = calculate_house_floor_area_demolished_by_year()

    house_change = calculate_house_change(households, database_manager.get_new_buildings_category_share())
    yearly_built_floor_area_house = calculate_yearly_built_floor_area_house(
        database_manager.get_building_category_floor_area(),
        demolition,
        house_change)

    scenario_r_house = pd.concat([
        demolition['demolition_change'],
        yearly_built_floor_area_house['yearly_new_house_floor_area'],
        calculate_new_buildings_accumulated(yearly_built_floor_area_house)['new_house_accumulated']],
        axis=1)

    regneark = pd.merge(left=facts, right=scenario_r_house, left_index=True, right_index=True)

    regneark = regneark.rename(axis=1, mapper={'population': 'Befolkning',
                                               'population_change': 'Befolkningsøkning',
                                               'household_size': 'Pers/Hushold',
                                               'households': 'Husholdninger',
                                               'households_change': 'Årlig endring antall boliger',
                                               'demolition_change': 'Årlig revet areal',
                                               'yearly_new_house_floor_area': 'Årlig nybygget areal',
                                               'new_house_accumulated': 'Nybygget småhus akkumulert'})
    display(regneark.transpose())


def run_notebook():
    print('# Nybygging formler')
    database_manager = DatabaseManager()
    print('### Andeler småhus/leiligheter')
    new_buildings_category_share = database_manager.get_new_buildings_category_share()
    display(new_buildings_category_share.head())
    print('### Husholdninger (scenario)')
    households = calculate_households(database_manager.get_construction_population())
    display(households[households.year == 2024])
    display(households.head())
    print('### Årlig endring antall småhus')
    yearly_change_small_house = calculate_house_change(households, new_buildings_category_share)
    display(yearly_change_small_house.head())
    print('### Årlig revet areal småhus')
    floor_area_demolished_by_year = calculate_house_floor_area_demolished_by_year()
    display(floor_area_demolished_by_year.head())
    display(floor_area_demolished_by_year.tail())
    print('### Årlig nybygget areal småhus')
    build_area_sum = calculate_yearly_built_floor_area_house(
        database_manager.get_building_category_floor_area(),
        floor_area_demolished_by_year,
        yearly_change_small_house)
    display(build_area_sum.head())
    padded_build_area_sum = pd.DataFrame(pd.concat([
        build_area_sum.loc['Small house'].loc[['2010', '2011']],
        pd.Series({y: np.float64(0) for y in range(2012, 2051)})]), columns=['area'])
    display(padded_build_area_sum.head())
    display(padded_build_area_sum.tail())
    print('## Nybygget småhus akkumulert')
    yearly_change_small_house = calculate_new_buildings_accumulated(yearly_change_small_house)
    display(yearly_change_small_house.head())
    display(yearly_change_small_house.head())


if __name__ == '__main__':
    main()
